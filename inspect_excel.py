import os
import openpyxl

base_dir = r"c:\WORK\00_INBOX\MAYIS BEYANLAR\MAYIS BEYANLAR"
excel_path = os.path.join(base_dir, "EXPORT.XLSX")

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
    
    # Read first 15 rows
    for r in range(1, min(16, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
        print(f"Row {r}: {row_vals}")
    wb.close()
else:
    print("EXPORT.XLSX not found at", excel_path)
