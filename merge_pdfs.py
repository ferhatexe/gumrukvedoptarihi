# -*- coding: utf-8 -*-
"""
Fatura ve Beyanname PDF'lerini eşleştirip birleştiren script.
Çıktı isimlendirme: FaturaNo1_FaturaNo2_FirmaIsmi_BeyanNo.pdf
Beyan PDF başa, fatura PDF'leri sırasıyla arkasına eklenir.
"""

import openpyxl
import os
import re
from PyPDF2 import PdfMerger
from collections import defaultdict

# Yollar
BASE_DIR = r"c:\WORK\00_INBOX\MAYIS BEYANLAR\MAYIS BEYANLAR"
EXCEL_PATH = os.path.join(BASE_DIR, "EXPORT.XLSX")
BEYANNAME_DIR = os.path.join(BASE_DIR, "BEYANNAME PDF")
FATURA_DIR = os.path.join(BASE_DIR, "İHRACAT MAYIS FATURALAR")
OUTPUT_DIR = os.path.join(BASE_DIR, "BİRLEŞTİRİLMİŞ")

# Çıktı klasörünü oluştur
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Excel'den verileri oku
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

# Fatura -> (firma, GCB) eşleştirmesi
# Aynı GCB numarasına sahip faturaları grupla
gcb_groups = defaultdict(lambda: {"firma": "", "faturalar": []})

for row in range(2, ws.max_row + 1):
    fatura_no = str(ws.cell(row=row, column=1).value).strip()
    firma = str(ws.cell(row=row, column=3).value).strip()
    gcb = str(ws.cell(row=row, column=9).value).strip()
    
    if not fatura_no or not gcb or fatura_no == "None" or gcb == "None":
        continue
    
    # Aynı fatura tekrar eklenmesin (BTC2026000002871 ve BTC2026000002610 gibi duplicate satırlar var)
    if fatura_no not in gcb_groups[gcb]["faturalar"]:
        gcb_groups[gcb]["faturalar"].append(fatura_no)
    gcb_groups[gcb]["firma"] = firma

print("=" * 80)
print("FATURA-BEYANNAME EŞLEŞTİRME RAPORU")
print("=" * 80)

# Mevcut dosyaları indexle
# Beyanname dosyaları: 26341200EX00137190_Beyanname.pdf
beyanname_files = {}
for f in os.listdir(BEYANNAME_DIR):
    if f.endswith(".pdf"):
        gcb_no = f.replace("_Beyanname.pdf", "")
        beyanname_files[gcb_no] = os.path.join(BEYANNAME_DIR, f)

print(f"\nToplam beyanname PDF: {len(beyanname_files)}")

# Fatura dosyaları: 1680036193_BT12026000000561005056bc-4acf-1fe1-91f7-2d92513b2f63.pdf
# Fatura no'yu dosya isminden çıkart
fatura_files = {}
for f in os.listdir(FATURA_DIR):
    if f.endswith(".pdf"):
        # Dosya isminden fatura numarasını çıkart
        # Format: 1680036193_BT12026000000561005056bc-...
        # veya: 1680036193_BTC2026000002387005056bc-...
        match = re.search(r'_(BT[C]?\d+?)005056', f)
        if match:
            fatura_no = match.group(1)
            fatura_files[fatura_no] = os.path.join(FATURA_DIR, f)

print(f"Toplam fatura PDF: {len(fatura_files)}")
print(f"\nBulunan fatura numaraları: {sorted(fatura_files.keys())}")

# Eşleştirme ve birleştirme
success_count = 0
error_count = 0
missing_beyan = []
missing_fatura = []

print("\n" + "-" * 80)
print("BİRLEŞTİRME İŞLEMLERİ")
print("-" * 80)

for gcb, info in sorted(gcb_groups.items()):
    firma = info["firma"]
    faturalar = sorted(info["faturalar"])
    
    print(f"\n[GCB] {gcb}")
    print(f"   Firma: {firma}")
    print(f"   Faturalar: {faturalar}")
    
    # Beyanname PDF kontrolu
    if gcb not in beyanname_files:
        print(f"   [HATA] BEYANNAME PDF BULUNAMADI: {gcb}")
        missing_beyan.append(gcb)
        error_count += 1
        continue
    
    # Fatura PDF kontrolu
    missing = []
    found_fatura_paths = []
    for fat in faturalar:
        if fat in fatura_files:
            found_fatura_paths.append((fat, fatura_files[fat]))
        else:
            missing.append(fat)
    
    if missing:
        print(f"   [UYARI] FATURA PDF BULUNAMADI: {missing}")
        missing_fatura.extend(missing)
    
    if not found_fatura_paths:
        print(f"   [HATA] HICBIR FATURA PDF BULUNAMADI, atlaniyor.")
        error_count += 1
        continue
    
    # Dosya ismini olustur
    # Format: FaturaNo1_FaturaNo2_FirmaIsmi_BeyanNo.pdf
    fatura_part = "_".join([fat for fat, _ in found_fatura_paths])
    # Firma ismindeki ozel karakterleri temizle
    clean_firma = firma.replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "").replace("?", "").replace('"', "").replace("<", "").replace(">", "").replace("|", "").replace("&", "and")
    output_name = f"{fatura_part}_{clean_firma}_{gcb}.pdf"
    output_path = os.path.join(OUTPUT_DIR, output_name)
    
    # PDF birlestirme: Once beyanname, sonra faturalar
    try:
        merger = PdfMerger()
        merger.append(beyanname_files[gcb])
        print(f"   [OK] Beyanname eklendi: {os.path.basename(beyanname_files[gcb])}")
        
        for fat, fat_path in found_fatura_paths:
            merger.append(fat_path)
            print(f"   [OK] Fatura eklendi: {fat}")
        
        merger.write(output_path)
        merger.close()
        
        print(f"   [CIKTI] {output_name}")
        success_count += 1
    except Exception as e:
        print(f"   [HATA] {e}")
        error_count += 1

print("\n" + "=" * 80)
print("ÖZET")
print("=" * 80)
print(f"Başarılı birleştirme: {success_count}")
print(f"Hata: {error_count}")

if missing_beyan:
    print(f"\nEksik beyanname PDF'leri ({len(missing_beyan)}):")
    for b in missing_beyan:
        print(f"  - {b}")

if missing_fatura:
    print(f"\nEksik fatura PDF'leri ({len(missing_fatura)}):")
    for f in missing_fatura:
        print(f"  - {f}")

print(f"\nÇıktı klasörü: {OUTPUT_DIR}")
