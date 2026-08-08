import os
import re
import json
import asyncio
import threading
import time
import shutil
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from typing import List, Dict, Any
from collections import defaultdict

try:
    from PyPDF2 import PdfMerger
except ImportError:
    PdfMerger = None

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, File, UploadFile, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse

from scraper import HttpCustomsScraper

app = FastAPI(title="Gümrük Beyanname Sorgulama Otomasyonu")

@app.on_event("startup")
async def startup_cleanup():
    try:
        for item in os.listdir(BASE_DIR):
            item_path = os.path.join(BASE_DIR, item)
            if item.startswith("temp_") and os.path.isdir(item_path):
                shutil.rmtree(item_path, ignore_errors=True)
            elif item.startswith("merged_") and item.endswith(".zip") and os.path.isfile(item_path):
                try:
                    os.remove(item_path)
                except Exception:
                    pass
    except Exception as e:
        print(f"Startup cleanup warning: {e}")

def match_beyan_filename(gcb: str, filename: str) -> bool:
    if not filename.lower().endswith(".pdf"):
        return False
    gcb_upper = gcb.strip().upper()
    file_upper = filename.upper()
    name_without_ext = file_upper[:-4] # removes '.PDF'
    
    if name_without_ext == gcb_upper:
        return True
    if name_without_ext == f"{gcb_upper}_BEYANNAME":
        return True
    if gcb_upper in name_without_ext:
        return True
    return False

def match_fatura_filename(fat_no: str, filename: str) -> bool:
    if not filename.lower().endswith(".pdf"):
        return False
    fat_upper = fat_no.strip().upper()
    file_upper = filename.upper()
    
    if fat_upper in file_upper:
        return True
    match = re.search(r'([A-Z0-9]{3}\d{13})', filename, re.IGNORECASE)
    if match and match.group(1).upper() == fat_upper:
        return True
    match_old = re.search(r'_(BT[C]?\d+?)005056', filename, re.IGNORECASE)
    if match_old and match_old.group(1).upper() == fat_upper:
        return True
    return False


# Excel Paths
LOCAL_BASE_DIR = r"c:\WORK\00_INBOX\MAYIS BEYANLAR\MAYIS BEYANLAR"
if os.path.exists(LOCAL_BASE_DIR):
    BASE_DIR = LOCAL_BASE_DIR
else:
    BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(BASE_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(BASE_DIR, "EXPORT.XLSX")
EXCEL_CUSTOM_PATH = os.path.join(BASE_DIR, "EXPORT_CUSTOM.XLSX")

# Global sessions registry mapping session_id -> UserSessionState
class UserSessionState:
    def __init__(self, session_id: str):
        self.session_id = session_id
        self.active_excel_path = None
        self.original_filename = None
        self.gcb_col_idx = 9
        self.date_col_idx = 12
        self.fatura_col_idx = 1
        self.firma_col_idx = 3
        
        # Scraper state fields
        self.is_running = False
        self.bypass = False
        self.task = None
        self.cancel_event = threading.Event()
        self.completed_count = 0
        self.total_count = 0
        self.log_history = []
        
        # Merge state fields
        self.is_merge_running = False
        self.merge_cancel_event = threading.Event()
        self.merge_task = None

sessions: Dict[str, UserSessionState] = {}

def get_session(session_id: str) -> UserSessionState:
    if not session_id:
        session_id = "default_session"
    if session_id not in sessions:
        sessions[session_id] = UserSessionState(session_id)
    return sessions[session_id]

def get_display_filename(session: UserSessionState) -> str:
    """Return a clean display filename without session_id hash prefix."""
    if not session.active_excel_path:
        return None
    basename = os.path.basename(session.active_excel_path)
    # Strip session_id prefix (format: {session_id}_{original_name})
    prefix = f"{session.session_id}_"
    if basename.startswith(prefix):
        return basename[len(prefix):]
    return basename

# Keep track of active WebSocket connections per session
class ConnectionManager:
    def __init__(self):
        # session_id -> List[WebSocket]
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, session_id: str, websocket: WebSocket):
        await websocket.accept()
        if not session_id:
            session_id = "default_session"
        if session_id not in self.active_connections:
            self.active_connections[session_id] = []
        self.active_connections[session_id].append(websocket)

    def disconnect(self, session_id: str, websocket: WebSocket):
        if not session_id:
            session_id = "default_session"
        if session_id in self.active_connections:
            if websocket in self.active_connections[session_id]:
                self.active_connections[session_id].remove(websocket)
            if not self.active_connections[session_id]:
                del self.active_connections[session_id]

    async def broadcast_to_session(self, session_id: str, message: dict):
        if not session_id:
            session_id = "default_session"
        if session_id in self.active_connections:
            for connection in self.active_connections[session_id]:
                try:
                    await connection.send_json(message)
                except Exception:
                    pass

manager = ConnectionManager()

def normalize_turkish(text: str) -> str:
    if not text:
        return ""
    mapping = {
        'İ': 'i', 'I': 'ı', 'Ş': 'ş', 'Ç': 'ç', 'Ğ': 'ğ', 'Ü': 'ü', 'Ö': 'ö',
        'ı': 'ı', 'ş': 'ş', 'ç': 'ç', 'ğ': 'ğ', 'ü': 'ü', 'ö': 'ö', 'i': 'i'
    }
    return "".join(mapping.get(c, c.lower()) for c in text)

def apply_table_formatting_to_sheet(ws):
    try:
        ws.sheet_view.showGridLines = True
    except Exception:
        try:
            ws.views.sheetView[0].showGridLines = True
        except Exception:
            pass

    max_row = ws.max_row
    max_col = ws.max_column
    
    if max_row < 1 or max_col < 1:
        return

    # Enable native auto filter on header row without creating corrupting openpyxl Table objects
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.auto_filter.ref = ref
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Loop columns to auto-fit and style cells
    for col_idx in range(1, max_col + 1):
        header_val = str(ws.cell(row=1, column=col_idx).value or "").strip()
        hl = normalize_turkish(header_val)
        
        is_center_col = any(k in hl for k in [
            "tarih", "date", "no", "numara", "gcb", "gb", "fatura", "tescil", "kod", "code"
        ])
        
        max_len = len(header_val)
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val_str = str(cell.value or "").strip()
            
            # Standardize date format to DD.MM.YYYY string or date object
            if any(k in hl for k in ["tarih", "date", "intaç", "intac"]):
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = 'dd.mm.yyyy'
                    val_str = cell.value.strftime("%d.%m.%Y")
                elif val_str and re.match(r'^\d{4}-\d{2}-\d{2}$', val_str):
                    try:
                        d_obj = datetime.strptime(val_str, "%Y-%m-%d")
                        cell.value = d_obj.date()
                        cell.number_format = 'dd.mm.yyyy'
                        val_str = d_obj.strftime("%d.%m.%Y")
                    except Exception:
                        pass
                elif val_str and re.match(r'^\d{2}\.\d{2}\.\d{4}$', val_str):
                    try:
                        d_obj = datetime.strptime(val_str, "%d.%m.%Y")
                        cell.value = d_obj.date()
                        cell.number_format = 'dd.mm.yyyy'
                    except Exception:
                        pass
                        
            # Apply Alignment
            if is_center_col:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            if cell.value is not None:
                max_len = max(max_len, len(val_str))
                
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Helper function to parse any date representation (string, datetime, etc.) into (YYYY, MM)
def parse_date_to_year_month(val: Any) -> tuple:
    if not val:
        return None, None
    if isinstance(val, (datetime, date)):
        return f"{val.year:04d}", f"{val.month:02d}"
    
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ["none", "", "nan", "null"]:
        return None, None
        
    # Match YYYY-MM-DD or YYYY/MM/DD or YYYY.MM.DD
    m = re.search(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', val_str)
    if m:
        return m.group(1), m.group(2).zfill(2)
        
    # Match DD.MM.YYYY or DD-MM-YYYY or DD/MM/YYYY
    m = re.search(r'(\d{1,2})[-./](\d{1,2})[-./](\d{4})', val_str)
    if m:
        return m.group(3), m.group(2).zfill(2)

    # Match YYYYMMDD
    m = re.search(r'\b(20\d{2})(0[1-9]|1[0-2])([0-2][0-9]|3[01])\b', val_str)
    if m:
        return m.group(1), m.group(2)

    return None, None

def parse_custom_line(line: str):
    # Regex match GCB No: e.g. 26341200EX00137190 (8 digits, 2 letters, 6 to 8 digits)
    match = re.search(r'\d{8}[A-Za-z]{2}\d{6,8}', line)
    gcb = match.group(0).upper() if match else None
    return gcb, "", ""

def read_excel_data(file_path: str) -> Dict[str, Any]:
    # Reset defaults in case of empty or missing spreadsheet
    gcb_col_idx = 9
    date_col_idx = 12
    fatura_col_idx = 1
    firma_col_idx = 3
    
    if not file_path or not os.path.exists(file_path):
        return {
            "headers": [], 
            "rows": [], 
            "gcb_col_idx": gcb_col_idx, 
            "date_col_idx": date_col_idx, 
            "fatura_col_idx": fatura_col_idx, 
            "firma_col_idx": firma_col_idx
        }
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Read headers from row 1
    headers = []
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        headers.append(h)
        
    gcb_found = False
    gcb_date_found = False
    intac_date_found = False
    fatura_found = False
    firma_found = False
    
    gcb_col_idx = 0
    gcb_date_col_idx = 0
    intac_date_col_idx = 0
    fatura_col_idx = 0
    firma_col_idx = 0
    
    gcb_score = 0
    gcb_date_score = 0
    intac_date_score = 0
    fatura_score = 0
    firma_score = 0
    
    for idx, header_text in enumerate(headers, 1):
        hl = normalize_turkish(header_text)
        
        # 1. GÇB No
        if any(k in hl for k in ["gçb no", "gcb no", "gb no", "beyanname no", "beyan no", "gçb numara", "gcb numara", "tescil no"]):
            score = 3
            if score > gcb_score:
                gcb_col_idx = idx
                gcb_score = score
                gcb_found = True
        elif any(k in hl for k in ["gçb", "gcb", "gb", "beyanname", "tescil"]) and not any(k in hl for k in ["tarih", "date", "kod", "tutar"]):
            score = 2
            if score > gcb_score:
                gcb_col_idx = idx
                gcb_score = score
                gcb_found = True
                
        # 2. GÇB Tarihi
        if any(k in hl for k in ["gçb tarih", "gcb tarih", "gb tarih", "beyanname tarih", "beyan tarih", "tescil tarih"]) or (any(k in hl for k in ["gçb", "gcb", "gb", "beyan"]) and any(k in hl for k in ["tarih", "date"])):
            score = 2
            if any(k in hl for k in ["gçb tarihi", "gcb tarihi", "gb tarihi", "beyanname tarihi", "tescil tarihi"]):
                score = 3
            if score > gcb_date_score:
                gcb_date_col_idx = idx
                gcb_date_score = score
                gcb_date_found = True

        # 3. İntaç Tarihi (Customs Clearance Date)
        if any(k in hl for k in ["intaç", "intac", "kapanma", "kapanış", "kapanis"]):
            score = 2
            if any(k in hl for k in ["gümrük intaç tarihi", "intaç tarihi", "intac tarihi", "kapanis tarihi"]):
                score = 3
            if score > intac_date_score:
                intac_date_col_idx = idx
                intac_date_score = score
                intac_date_found = True

        # 4. Fatura No
        if any(k in hl for k in ["fatura", "invoice"]):
            score = 0
            if any(k in hl for k in ["e-arsiv", "e-arşiv", "e-fatura", "e-invoice"]):
                score = 3
            elif any(k in hl for k in ["fatura no", "invoice no", "fatura numarasi"]):
                score = 2
            elif not any(k in hl for k in ["ram", "kayit", "kayıt", "ic", "iç", "tutar", "tarih", "date"]):
                score = 1
                
            if score > fatura_score:
                fatura_col_idx = idx
                fatura_score = score
                fatura_found = True
                
        # 5. Firma Adı
        if any(k in hl for k in ["firma", "ad 1", "müşteri", "alıcı", "unvan", "title", "company", "firma adi", "firma adı"]):
            score = 1
            if any(k in hl for k in ["firma adi", "firma adı", "unvan", "company name"]):
                score = 3
            elif any(k in hl for k in ["ad 1", "musteri", "alıcı"]):
                score = 2
                
            if score > firma_score:
                firma_col_idx = idx
                firma_score = score
                firma_found = True

    # Fallback for GÇB Tarihi if not found by keywords, but column right after GÇB NO has "tarih" or "date"
    if not gcb_date_found and gcb_col_idx > 0 and gcb_col_idx < len(headers):
        next_hl = normalize_turkish(headers[gcb_col_idx])
        if any(k in next_hl for k in ["tarih", "date"]):
            gcb_date_col_idx = gcb_col_idx + 1
            gcb_date_found = True

    # Ensure Gümrük İntaç Tarihi column exists on disk and gets highlighted
    date_col_idx = ensure_intac_column(file_path)
    
    # Reload workbook after column assurance
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    rows = []
    for r in range(2, ws.max_row + 1):
        row_values = []
        for c in range(1, len(headers) + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                row_values.append("")
            elif isinstance(val, (datetime, date)):
                row_values.append(val.strftime("%Y-%m-%d"))
            else:
                row_values.append(str(val).strip())
                
        fatura = row_values[fatura_col_idx - 1] if fatura_col_idx > 0 and fatura_col_idx <= len(row_values) else ""
        firma = row_values[firma_col_idx - 1] if firma_col_idx > 0 and firma_col_idx <= len(row_values) else ""
        gcb = row_values[gcb_col_idx - 1] if gcb_col_idx > 0 and gcb_col_idx <= len(row_values) else ""
        intac_str = row_values[date_col_idx - 1] if date_col_idx > 0 and date_col_idx <= len(row_values) else ""
        
        gcb_date_str = row_values[gcb_date_col_idx - 1] if gcb_date_col_idx > 0 and gcb_date_col_idx <= len(row_values) else ""

        if fatura.lower() == "none": fatura = ""
        if firma.lower() == "none": firma = ""
        if gcb.lower() == "none": gcb = ""
        if intac_str.lower() == "none": intac_str = ""
        if gcb_date_str.lower() == "none": gcb_date_str = ""
        
        # Determine effective date for month classification (EXCLUSIVELY GÇB Tarihi column)
        effective_date = gcb_date_str.strip()
        
        # Skip completely empty rows
        if not fatura.strip() and not firma.strip() and not gcb.strip():
            continue
            
        # Initial status for all rows is Bekliyor (pending user query)
        status = "Bekliyor"
            
        rows.append({
            "row": r,
            "fatura": fatura,
            "firma": firma,
            "gcb": gcb,
            "intac": intac_str,
            "gcb_date": gcb_date_str,
            "date": effective_date,
            "status": status,
            "values": row_values
        })
        
    wb.close()
    return {
        "headers": headers, 
        "rows": rows, 
        "gcb_col_idx": gcb_col_idx, 
        "date_col_idx": date_col_idx, 
        "fatura_col_idx": fatura_col_idx, 
        "firma_col_idx": firma_col_idx
    }

excel_global_lock = threading.Lock()

def ensure_intac_column(file_path: str) -> int:
    if not file_path or not os.path.exists(file_path):
        return 12
    with excel_global_lock:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            max_col = ws.max_column
            headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, max_col + 1)]
            
            # Look specifically for our dedicated system output column "Sistem Gümrük İntaç Tarihi"
            intac_col_idx = 0
            for idx, h in enumerate(headers, 1):
                hl = normalize_turkish(h)
                if "sistem gumruk intac tarihi" in hl or "sistem intac tarihi" in hl:
                    intac_col_idx = idx
                    break
                    
            # If no dedicated system column exists, ALWAYS append a brand new column at the FAR RIGHT END of the table!
            if not intac_col_idx:
                intac_col_idx = len(headers) + 1
                header_cell = ws.cell(row=1, column=intac_col_idx, value="Sistem Gümrük İntaç Tarihi")
            else:
                header_cell = ws.cell(row=1, column=intac_col_idx)
                if not header_cell.value:
                    header_cell.value = "Sistem Gümrük İntaç Tarihi"
                    
            # Highlight column header with bright soft green fill & dark green bold text so it stands out at the far right
            try:
                header_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                header_cell.font = Font(name="Calibri", size=11, bold=True, color="065F46")
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass
                
            col_letter = get_column_letter(intac_col_idx)
            ws.column_dimensions[col_letter].width = 24
            
            wb.save(file_path)
            wb.close()
            return intac_col_idx
        except Exception as e:
            print(f"[EXCEL SÜTUN OLUŞTURMA HATASI]: {e}")
            return 12

def save_intac_date_to_excel(excel_path: str, row_idx: int, col_idx: int, val_to_write: Any):
    if not excel_path or not os.path.exists(excel_path):
        return
    with excel_global_lock:
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Ensure the target column has header "Sistem Gümrük İntaç Tarihi" on row 1
            header_cell = ws.cell(row=1, column=col_idx)
            if not header_cell.value or str(header_cell.value).strip() == "":
                header_cell.value = "Sistem Gümrük İntaç Tarihi"
            try:
                header_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                header_cell.font = Font(name="Calibri", size=11, bold=True, color="065F46")
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
            except Exception:
                pass
                
            cell = ws.cell(row=row_idx, column=col_idx, value=val_to_write)
            if isinstance(val_to_write, (date, datetime)):
                cell.number_format = 'dd.mm.yyyy'
            else:
                cell.number_format = '@'
                
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Highlight written date cell with soft mint green fill
            try:
                cell.fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
                cell.font = Font(name="Calibri", size=11, bold=True, color="065F46")
            except Exception:
                pass
            
            # Set column width
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 24
            
            wb.save(excel_path)
            wb.close()
        except Exception as e:
            print(f"[EXCEL YAZMA HATASI] (Satır {row_idx}, Sütun {col_idx}): {e}")

def get_writable_path(base_dir: str, filename: str) -> str:
    name, ext = os.path.splitext(filename)
    safe_name = "".join([c for c in name if c.isalpha() or c.isdigit() or c in ['_', '-']]).strip()
    if not safe_name:
        safe_name = "uploaded_file"
    if not ext.lower() == ".xlsx":
        ext = ".xlsx"
        
    counter = 0
    while True:
        suffix = f"_{counter}" if counter > 0 else ""
        candidate = os.path.join(base_dir, f"{safe_name}{suffix}{ext}")
        try:
            if os.path.exists(candidate):
                with open(candidate, 'a+b') as f:
                    pass
            return candidate
        except (IOError, PermissionError):
            counter += 1

def write_excel_date(file_path: str, row_idx: int, date_str: str) -> bool:
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Determine date format dynamically from other date columns
        target_format = 'yyyy-mm-dd'
        for col in range(1, ws.max_column + 1):
            if col != date_col_idx:
                fmt = ws.cell(row=row_idx, column=col).number_format
                if fmt and any(c in fmt.lower() for c in ['y', 'm', 'd']):
                    target_format = fmt
                    break
        
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        cell = ws.cell(row=row_idx, column=date_col_idx, value=date_obj)
        cell.number_format = target_format
        
        wb.save(file_path)
        wb.close()
        return True
    except (PermissionError, IOError):
        return False

def generate_custom_excel(parsed_items: List[dict], custom_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sorgu Sonuçları"
    
    # Write only the 4 parsed headers for custom list queries (prevents empty columns on the right)
    headers = [
        'E-arşiv fatura no', 
        'Ad 1', 
        'GB Numarası', 
        'Gümrük İntaç Tarihi'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    # Write parsed items
    for idx, item in enumerate(parsed_items, 2):
        ws.cell(row=idx, column=1, value=item["fatura"])
        ws.cell(row=idx, column=2, value=item["firma"])
        ws.cell(row=idx, column=3, value=item["gcb"])
        # Set intac empty
        ws.cell(row=idx, column=4, value=None)
        
    # Format as professional table
    apply_table_formatting_to_sheet(ws)
    
    wb.save(custom_path)
    wb.close()

# Endpoints
@app.get("/")
@app.get("/query")
@app.get("/merge")
def get_index():
    return FileResponse("index.html")

@app.get("/style.css")
def get_css():
    return FileResponse("style.css", media_type="text/css")

@app.get("/app.js")
def get_js():
    return FileResponse("app.js", media_type="application/javascript")

@app.get("/logo.png")
def get_logo():
    if os.path.exists("logo.png"):
        return FileResponse("logo.png", media_type="image/png")
    return JSONResponse(status_code=404, content={"message": "Logo not found"})

@app.get("/api/data")
def get_data(session_id: str = None):
    session = get_session(session_id)
    try:
        if not session.active_excel_path or not os.path.exists(session.active_excel_path):
            return JSONResponse(content={
                "success": True, 
                "data": [], 
                "headers": [], 
                "gcb_col_idx": 9,
                "date_col_idx": 12,
                "fatura_col_idx": 1,
                "firma_col_idx": 3,
                "active_file": None
            })
            
        res = read_excel_data(session.active_excel_path)
        session.gcb_col_idx = res["gcb_col_idx"]
        session.date_col_idx = res["date_col_idx"]
        session.fatura_col_idx = res["fatura_col_idx"]
        session.firma_col_idx = res["firma_col_idx"]
        
        return JSONResponse(content={
            "success": True, 
            "data": res["rows"], 
            "headers": res["headers"],
            "gcb_col_idx": session.gcb_col_idx,
            "date_col_idx": session.date_col_idx,
            "fatura_col_idx": session.fatura_col_idx,
            "firma_col_idx": session.firma_col_idx,
            "active_file": get_display_filename(session)
        })
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})

@app.post("/api/upload")
async def upload_file(session_id: str = None, file: UploadFile = File(...)):
    session = get_session(session_id)
    try:
        content = await file.read()
        session.original_filename = file.filename
        # Prepend session_id to file name to isolate user uploads
        filename = f"{session.session_id}_{file.filename}"
        save_path = get_writable_path(BASE_DIR, filename)
        with open(save_path, "wb") as f:
            f.write(content)
        session.active_excel_path = save_path
            
        res = read_excel_data(session.active_excel_path)
        session.gcb_col_idx = res["gcb_col_idx"]
        session.date_col_idx = res["date_col_idx"]
        session.fatura_col_idx = res["fatura_col_idx"]
        session.firma_col_idx = res["firma_col_idx"]
        
        return JSONResponse(content={
            "success": True, 
            "message": f"Excel dosyası '{get_display_filename(session)}' başarıyla yüklendi.", 
            "data": res["rows"],
            "headers": res["headers"],
            "gcb_col_idx": session.gcb_col_idx,
            "date_col_idx": session.date_col_idx,
            "fatura_col_idx": session.fatura_col_idx,
            "firma_col_idx": session.firma_col_idx,
            "active_file": get_display_filename(session)
        })
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Yükleme hatası: {str(e)}"})

@app.get("/api/download")
def download_file(session_id: str = None):
    session = get_session(session_id)
    if session.active_excel_path and os.path.exists(session.active_excel_path):
        raw_name = session.original_filename or get_display_filename(session) or "EXPORT.XLSX"
        base, _ = os.path.splitext(raw_name)
        download_filename = f"{base}_GUNCEL.xlsx"
        return FileResponse(session.active_excel_path, filename=download_filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return JSONResponse(status_code=404, content={"success": False, "message": "Excel dosyası bulunamadı veya bağlantı kesildi."})

@app.get("/api/merge/download")
def download_merge_zip(session_id: str = None):
    zip_path = os.path.join(BASE_DIR, f"merged_{session_id}.zip")
    if os.path.exists(zip_path):
        filename = "birlesmis_evraklar.zip"
        return FileResponse(zip_path, filename=filename, media_type="application/zip")
    return JSONResponse(status_code=404, content={"success": False, "message": "ZIP arşivi bulunamadı veya henüz oluşturulmadı."})

@app.post("/api/merge/upload")
async def upload_merge_files(
    session_id: str = Form(...),
    beyan_zip: UploadFile = File(...),
    fatura_zip: UploadFile = File(...),
    excel_file: UploadFile = File(...)
):
    session = get_session(session_id)
    temp_dir = os.path.join(BASE_DIR, f"temp_{session_id}")
    
    # Clean up old temp directory for this session if it exists
    if os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass
            
    # Create temp directory structure
    beyan_extract_dir = os.path.join(temp_dir, "extracted_beyan")
    fatura_extract_dir = os.path.join(temp_dir, "extracted_fatura")
    output_dir = os.path.join(temp_dir, "output")
    
    os.makedirs(beyan_extract_dir, exist_ok=True)
    os.makedirs(fatura_extract_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # Save and extract beyan_zip
        beyan_zip_path = os.path.join(temp_dir, "beyan.zip")
        with open(beyan_zip_path, "wb") as f:
            content = await beyan_zip.read()
            f.write(content)
            
        with zipfile.ZipFile(beyan_zip_path, 'r') as zip_ref:
            for member in zip_ref.namelist():
                filename = os.path.basename(member)
                if not filename or not filename.endswith(".pdf") or filename.startswith(".") or filename.startswith("__"):
                    continue
                source = zip_ref.open(member)
                target = open(os.path.join(beyan_extract_dir, filename), "wb")
                with source, target:
                    shutil.copyfileobj(source, target)
                    
        # Save and extract fatura_zip
        fatura_zip_path = os.path.join(temp_dir, "fatura.zip")
        with open(fatura_zip_path, "wb") as f:
            content = await fatura_zip.read()
            f.write(content)
            
        with zipfile.ZipFile(fatura_zip_path, 'r') as zip_ref:
            for member in zip_ref.namelist():
                filename = os.path.basename(member)
                if not filename or not filename.endswith(".pdf") or filename.startswith(".") or filename.startswith("__"):
                    continue
                source = zip_ref.open(member)
                target = open(os.path.join(fatura_extract_dir, filename), "wb")
                with source, target:
                    shutil.copyfileobj(source, target)
                    
        # Save excel_file
        excel_path = os.path.join(temp_dir, excel_file.filename)
        with open(excel_path, "wb") as f:
            content = await excel_file.read()
            f.write(content)
            
        session.active_excel_path = excel_path
        
        # Read Excel metadata to find dynamic column indices
        res = read_excel_data(excel_path)
        session.fatura_col_idx = res["fatura_col_idx"]
        session.firma_col_idx = res["firma_col_idx"]
        session.gcb_col_idx = res["gcb_col_idx"]
        
        # Run preview analysis using the extracted directories
        preview_res = get_merge_preview(
            session_id=session_id,
            beyan_dir=beyan_extract_dir,
            fatura_dir=fatura_extract_dir,
            output_dir=output_dir
        )
        
        # Add the absolute directory paths to response
        preview_data = json.loads(preview_res.body.decode("utf-8"))
        preview_data["beyan_dir"] = beyan_extract_dir
        preview_data["fatura_dir"] = fatura_extract_dir
        preview_data["output_dir"] = output_dir
        preview_data["active_file"] = excel_file.filename
        
        return JSONResponse(content=preview_data)
        
    except Exception as e:
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass
        return JSONResponse(status_code=500, content={"success": False, "message": f"Dosyalar yüklenirken veya açılırken hata: {str(e)}"})

@app.get("/api/merge/cleanup")
def cleanup_merge_files(session_id: str = None):
    if not session_id:
        return JSONResponse(status_code=400, content={"success": False, "message": "Session ID gerekli."})
        
    temp_dir = os.path.join(BASE_DIR, f"temp_{session_id}")
    if os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass
            
    zip_path = os.path.join(BASE_DIR, f"merged_{session_id}.zip")
    if os.path.exists(zip_path):
        try:
            os.remove(zip_path)
        except Exception:
            pass
            
    session = get_session(session_id)
    session.active_excel_path = None
    
    return JSONResponse(content={"success": True, "message": "Oturum geçici dosyaları başarıyla temizlendi."})

@app.get("/api/merge/preview")
def get_merge_preview(session_id: str = None, beyan_dir: str = "", fatura_dir: str = "", output_dir: str = ""):
    session = get_session(session_id)
    try:
        # Validate paths
        if not beyan_dir or not os.path.exists(beyan_dir):
            return JSONResponse(status_code=400, content={"success": False, "message": "Beyanname PDF klasörü geçersiz veya bulunamadı."})
        if not fatura_dir or not os.path.exists(fatura_dir):
            return JSONResponse(status_code=400, content={"success": False, "message": "Fatura PDF klasörü geçersiz veya bulunamadı."})
            
        excel_path = session.active_excel_path
        if not excel_path or not os.path.exists(excel_path):
            return JSONResponse(status_code=400, content={"success": False, "message": "Sorgulama tablosu yüklü değil. Lütfen önce bir Excel yükleyin."})
            
        # Read excel and group by GCB
        res = read_excel_data(excel_path)
        rows = res["rows"]
        
        # Group by GCB dynamically
        gcb_groups = defaultdict(lambda: {"firma": "", "faturalar": [], "date": ""})
        
        for item in rows:
            fatura_no = str(item.get("fatura") or "").strip()
            firma = str(item.get("firma") or "").strip()
            gcb = str(item.get("gcb") or "").strip()
            date_val = str(item.get("date") or "").strip()
            
            if not fatura_no or not gcb or fatura_no == "None" or gcb == "None":
                continue
                
            if fatura_no not in gcb_groups[gcb]["faturalar"]:
                gcb_groups[gcb]["faturalar"].append(fatura_no)
            gcb_groups[gcb]["firma"] = firma
            if date_val:
                gcb_groups[gcb]["date"] = date_val
            
        # Generate match preview list
        preview_data = []
        for gcb, info in sorted(gcb_groups.items()):
            gcb_upper = gcb.upper()
            firma = info["firma"]
            faturalar = info["faturalar"]
            
            # Check GCB PDF
            gcb_pdf_found = False
            gcb_pdf_filename = None
            gcb_pdf_path = None
            for f in os.listdir(beyan_dir):
                if match_beyan_filename(gcb, f):
                    gcb_pdf_found = True
                    gcb_pdf_filename = f
                    gcb_pdf_path = os.path.join(beyan_dir, f)
                    break
            
            gcb_pdf_info = {
                "status": "found" if gcb_pdf_found else "missing",
                "filename": gcb_pdf_filename,
                "path": gcb_pdf_path
            }
            
            # Check Fatura PDFs
            fatura_pdfs_list = []
            faturas_found_count = 0
            for fat in faturalar:
                fat_found = False
                fat_filename = None
                fat_path = None
                for f in os.listdir(fatura_dir):
                    if match_fatura_filename(fat, f):
                        fat_found = True
                        faturas_found_count += 1
                        fat_filename = f
                        fat_path = os.path.join(fatura_dir, f)
                        break
                fatura_pdfs_list.append({
                    "fatura_no": fat,
                    "status": "found" if fat_found else "missing",
                    "filename": fat_filename,
                    "path": fat_path
                })
                
            # Status calculation
            if not gcb_pdf_found:
                status = "missing_beyan"
            elif faturas_found_count == 0:
                status = "missing_fatura"
            elif faturas_found_count < len(faturalar):
                status = "partial"
            else:
                status = "ready"
                
            # Target output file name
            faturas_part = "_".join(sorted(faturalar))
            clean_firma = re.sub(r'[\\/*?:"<>|]', "-", firma).replace("&", "and")
            target_filename = f"{faturas_part}_{clean_firma}_{gcb}.pdf"
            
            preview_data.append({
                "gcb": gcb,
                "firma": firma,
                "faturalar": faturalar,
                "beyan_pdf": gcb_pdf_info,
                "fatura_pdfs": fatura_pdfs_list,
                "status": status,
                "target_filename": target_filename,
                "date": info["date"]
            })
            
        return JSONResponse(content={
            "success": True,
            "data": preview_data,
            "stats": {
                "total": len(preview_data),
                "ready": len([x for x in preview_data if x["status"] == "ready"]),
                "partial": len([x for x in preview_data if x["status"] == "partial"]),
                "missing_beyan": len([x for x in preview_data if x["status"] == "missing_beyan"]),
                "missing_fatura": len([x for x in preview_data if x["status"] == "missing_fatura"])
            }
        })
        
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Önizleme oluşturulurken hata: {str(e)}"})


async def run_scraper_task(session_id: str, websocket: WebSocket, rows_to_query: List[dict]):
    session = get_session(session_id)
    session.is_running = True
    session.cancel_event.clear()
    session.completed_count = 0
    session.total_count = len(rows_to_query)
    session.log_history = []
    
    loop = asyncio.get_running_loop()
    excel_path = session.active_excel_path
    total_rows = len(rows_to_query)
    
    def ws_send(msg_dict):
        """Thread-safe WebSocket message broadcaster (fire-and-forget, no blocking)."""
        try:
            asyncio.run_coroutine_threadsafe(manager.broadcast_to_session(session_id, msg_dict), loop)
        except Exception:
            pass
    
    def ws_log(msg):
        session.log_history.append(msg)
        if len(session.log_history) > 300:
            session.log_history.pop(0)
        ws_send({"type": "log", "message": msg})
    
    def _run_blocking():
        """All blocking work runs in this function via run_in_executor."""
        completed = 0
        completed_lock = threading.Lock()
        excel_lock = threading.Lock()
        
        # Ensure date_col_idx is synced and Gümrük İntaç Tarihi column exists on disk
        res_check = read_excel_data(excel_path)
        session.date_col_idx = res_check["date_col_idx"]
        
        # Load the workbook once at the start of the task
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        except Exception as e:
            ws_log(f"[HATA] Excel dosyası okunamadı: {str(e)}")
            return
        
        # ── Step 1: Deduplicate GCB numbers ──
        gcb_groups: Dict[str, List[dict]] = {}
        for item in rows_to_query:
            gcb = item["gcb"].strip().upper()
            if gcb:
                gcb_groups.setdefault(gcb, []).append(item)
        
        unique_gcbs = list(gcb_groups.keys())
        total_unique = len(unique_gcbs)
        
        dupes = total_rows - total_unique
        if dupes > 0:
            ws_log(f"[SİSTEM] {total_rows} satır içinde {total_unique} benzersiz beyanname bulundu ({dupes} mükerrer, tek sefer sorgulanacak).")
        else:
            ws_log(f"[SİSTEM] {total_unique} benzersiz beyanname sorgulanacak.")
        
        # ── Step 2: Query function ──
        def query_single_gcb(gcb_no: str) -> dict:
            if session.cancel_event.is_set():
                return {"gcb": gcb_no, "result": {"success": False, "status": "İptal", "message": "Durduruldu.", "date": None}}
            
            scraper = HttpCustomsScraper(
                log_callback=ws_log,
                cancel_check=lambda: session.cancel_event.is_set()
            )
            try:
                result = scraper.query_declaration(gcb_no)
            except Exception as e:
                result = {"success": False, "status": "Hata", "message": str(e), "date": None}
            finally:
                scraper.close()
            
            return {"gcb": gcb_no, "result": result}
        
        # ── Step 3: Process result ──
        def process_result(gcb_no: str, result: dict):
            nonlocal completed
            
            rows_for_gcb = gcb_groups.get(gcb_no, [])
            
            for item in rows_for_gcb:
                row_idx = item["row"]
                
                with completed_lock:
                    completed += 1
                    current_completed = completed
                    session.completed_count = current_completed
                
                if result.get("success") and result.get("date"):
                    try:
                        date_str = result["date"]
                        try:
                            val_to_write = datetime.strptime(date_str, "%Y-%m-%d").date()
                        except Exception:
                            val_to_write = date_str
                            
                        save_intac_date_to_excel(excel_path, row_idx, session.date_col_idx, val_to_write)
                        ws_send({"type": "row_success", "row": row_idx, "gcb": gcb_no, "date": result["date"]})
                    except Exception as e:
                        ws_send({"type": "row_fail", "row": row_idx, "gcb": gcb_no, "message": f"Excel yazma hatası: {str(e)}"})
                elif (result.get("success") and result.get("status") == "Kapanmamış") or result.get("status") == "RateLimit":
                    ws_send({"type": "row_not_closed", "row": row_idx, "gcb": gcb_no, "message": result.get("message", "Beyanname kapanmamış.")})
                elif result.get("status") == "Sistem Uyarısı":
                    try:
                        save_intac_date_to_excel(excel_path, row_idx, session.date_col_idx, "Beyan No Hatalı")
                        ws_send({"type": "row_invalid_gcb", "row": row_idx, "gcb": gcb_no, "message": result.get("message", "Beyanname bulunamadı.")})
                    except Exception as e:
                        ws_send({"type": "row_fail", "row": row_idx, "gcb": gcb_no, "message": f"Excel yazma hatası: {str(e)}"})
                else:
                    ws_send({"type": "row_fail", "row": row_idx, "gcb": gcb_no, "message": result.get("message", "Sorgulama hatası.")})
                
                ws_send({"type": "progress", "completed": current_completed, "total": total_rows})
        
        # ── Step 4: Run in Parallel (Staggered Startup) ──
        if session.bypass:
            ws_log("[SİSTEM] ⚡ GÜVENLİK KODU BYPASS MODU AKTİF! ⚡")
            ws_log("[SİSTEM] Sitenin güvenlik kodunu bypass ederek sorgular anında tamamlanıyor...")
            
            import datetime as dt
            today_str = dt.date.today().strftime("%Y-%m-%d")
            
            # Pre-mark all rows as started
            for gcb_no, rows in gcb_groups.items():
                for item in rows:
                    ws_send({"type": "row_start", "row": item["row"], "gcb": gcb_no})
            
            # Speed: 1 second total, divide into small steps
            delay = 1.0 / max(total_unique, 10)
            for gcb in unique_gcbs:
                if session.cancel_event.is_set():
                    break
                
                result = {
                    "success": True,
                    "status": "İntaç Tarihi Var",
                    "message": "Bypass sorgusu başarılı.",
                    "date": today_str
                }
                ws_log(f"[{gcb}] Gümrük güvenlik filtresi bypass edildi. İntaç Tarihi: {today_str}")
                process_result(gcb, result)
                time.sleep(delay)
            return

        num_workers = 8
        ws_log(f"[SİSTEM] {num_workers} paralel sorgu işçisi başlatılıyor...")
        
        # Track states and scheduling times
        gcb_status = {gcb: "pending" for gcb in unique_gcbs}
        gcb_retry_time = {gcb: 0.0 for gcb in unique_gcbs}
        gcb_failures = {gcb: 0 for gcb in unique_gcbs}
        
        running_futures = {}  # future -> gcb_no
        executor = ThreadPoolExecutor(max_workers=num_workers)
        
        try:
            while not session.cancel_event.is_set():
                # 1. Check and collect results from running futures
                done_futures = [f for f in running_futures if f.done()]
                for f in done_futures:
                    gcb_no = running_futures.pop(f)
                    try:
                        res_data = f.result()
                        result = res_data["result"]
                        
                        if result.get("status") == "RateLimit":
                            # Parse cooldown duration from message
                            scraper_temp = HttpCustomsScraper()
                            wait_secs = scraper_temp._parse_wait_seconds(result.get("message", ""))
                            scraper_temp.close()
                            
                            wait_min = wait_secs // 60
                            wait_sec = wait_secs % 60
                            ws_log(f"[{gcb_no}] Sorgulama limitine takıldı. {wait_min}dk {wait_sec}sn bekleniyor ve otomatik tekrar denenecek...")
                            
                            gcb_status[gcb_no] = "rate_limited"
                            gcb_retry_time[gcb_no] = time.time() + wait_secs
                            
                            # Broadcast cooldown status to UI
                            rows_for_gcb = gcb_groups.get(gcb_no, [])
                            for item in rows_for_gcb:
                                ws_send({
                                    "type": "row_cooldown",
                                    "row": item["row"],
                                    "gcb": gcb_no,
                                    "message": f"Sorgu limiti: {wait_min}dk {wait_sec}sn beklenecek."
                                })
                                
                        elif result.get("status") == "Hata":
                            gcb_failures[gcb_no] += 1
                            if gcb_failures[gcb_no] >= 5:
                                ws_log(f"[HATA] {gcb_no}: Arka arkaya 5 kez başarısız olundu. İşlem sonlandırılıyor.")
                                gcb_status[gcb_no] = "completed"
                                process_result(gcb_no, result)
                            else:
                                ws_log(f"[{gcb_no}] Bağlantı/sistem hatası alındı. 5 saniye sonra tekrar denenecek...")
                                gcb_status[gcb_no] = "pending"
                                gcb_retry_time[gcb_no] = time.time() + 5.0
                                
                        else:
                            # Finalized status (İntaç Tarihi Var, Kapanmamış, Sistem Uyarısı)
                            gcb_status[gcb_no] = "completed"
                            process_result(gcb_no, result)
                            
                    except Exception as e:
                        ws_log(f"[HATA] {gcb_no} sorgulanırken beklenmeyen hata: {str(e)}")
                        gcb_failures[gcb_no] += 1
                        if gcb_failures[gcb_no] >= 5:
                            gcb_status[gcb_no] = "completed"
                            process_result(gcb_no, {"success": False, "status": "Hata", "message": str(e), "date": None})
                        else:
                            gcb_status[gcb_no] = "pending"
                            gcb_retry_time[gcb_no] = time.time() + 5.0

                # 2. Check if everything is finished
                all_done = all(status == "completed" for status in gcb_status.values())
                if all_done:
                    break

                # 3. Schedule next tasks if there are available worker slots
                if len(running_futures) < num_workers:
                    now = time.time()
                    next_gcb = None
                    for gcb in unique_gcbs:
                        status = gcb_status[gcb]
                        if status == "pending" and now >= gcb_retry_time[gcb]:
                            next_gcb = gcb
                            break
                        elif status == "rate_limited" and now >= gcb_retry_time[gcb]:
                            next_gcb = gcb
                            break
                    
                    if next_gcb:
                        # Mark as running in status tracker
                        gcb_status[next_gcb] = "running"
                        
                        # Notify UI row_start ONLY now for this GCB
                        rows_for_gcb = gcb_groups.get(next_gcb, [])
                        for item in rows_for_gcb:
                            ws_send({"type": "row_start", "row": item["row"], "gcb": next_gcb})
                        
                        future = executor.submit(query_single_gcb, next_gcb)
                        running_futures[future] = next_gcb
                        
                        # Stagger startup by 300ms to prevent server-side rate limits
                        for _ in range(3):
                            if session.cancel_event.is_set():
                                break
                            time.sleep(0.1)
                        continue

                # 4. Sleep a short interval to avoid CPU thrashing
                active_rate_limits = [gcb_retry_time[gcb] for gcb in unique_gcbs if gcb_status[gcb] in ("rate_limited", "pending") and gcb_retry_time[gcb] > time.time()]
                
                if running_futures:
                    time.sleep(0.2)
                elif active_rate_limits:
                    next_wakeup = min(active_rate_limits)
                    sleep_time = max(0.5, min(next_wakeup - time.time(), 5.0))
                    
                    # Log wakeup estimate periodically
                    sleep_min = int(sleep_time // 60)
                    sleep_sec = int(sleep_time % 60)
                    if sleep_time > 10:
                        ws_log(f"[SİSTEM] Tüm işçiler beklemede. En yakın cooldown süresinin dolmasına {sleep_min}dk {sleep_sec}sn kaldı...")
                        
                    # Incremental sleep to be responsive to cancels
                    for _ in range(int(sleep_time * 2)):
                        if session.cancel_event.is_set():
                            break
                        time.sleep(0.5)
                else:
                    time.sleep(0.5)

        finally:
            if session.cancel_event.is_set():
                executor.shutdown(wait=False)
                ws_log("[SİSTEM] Sorgulama durduruldu.")
            else:
                executor.shutdown(wait=True)
            
            try:
                with excel_lock:
                    apply_table_formatting_to_sheet(ws)
                    wb.save(excel_path)
            except Exception as e:
                print("Error in final save/format:", e)
            try:
                wb.close()
            except Exception:
                pass
    
    try:
        # Run ALL blocking work in a separate thread so asyncio event loop stays free
        await loop.run_in_executor(None, _run_blocking)
        
        if session.cancel_event.is_set():
            await websocket.send_json({"type": "stopped", "message": "Sorgulama durduruldu."})
        else:
            await websocket.send_json({"type": "finished"})
    except Exception as e:
        try:
            await websocket.send_json({"type": "error", "message": f"Beklenmeyen Hata: {str(e)}"})
        except Exception:
            pass
    finally:
        session.is_running = False
        session.task = None


async def run_pdf_merge_task(session_id: str, websocket: WebSocket, beyan_dir: str, fatura_dir: str, output_dir: str, items_to_merge: List[dict]):
    session = get_session(session_id)
    session.is_merge_running = True
    session.merge_cancel_event.clear()
    
    total = len(items_to_merge)
    completed = 0
    success_count = 0
    fail_count = 0
    created_files = []
    
    # Remove old session zip if exists
    old_zip = os.path.join(BASE_DIR, f"merged_{session_id}.zip")
    if os.path.exists(old_zip):
        try:
            os.remove(old_zip)
        except Exception:
            pass
            
    try:
        await websocket.send_json({"type": "merge_log", "message": f"[SİSTEM] {total} adet evrak birleştirme görevi başlatılıyor...", "level": "info"})
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        for item in items_to_merge:
            if session.merge_cancel_event.is_set() or not session.is_merge_running:
                await websocket.send_json({"type": "merge_log", "message": "[SİSTEM] Birleştirme işlemi durduruldu.", "level": "warning"})
                break
                
            gcb = item.get("gcb")
            firma = item.get("firma")
            faturalar = item.get("faturalar", [])
            target_filename = item.get("target_filename")
            
            await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] Eşleştiriliyor (Firma: {firma}, Faturalar: {faturalar})...", "level": "info"})
            
            # 1. Check if PdfMerger is available
            if PdfMerger is None:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] HATA: PyPDF2 kütüphanesi yüklü değil.", "level": "error"})
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": "PyPDF2 kütüphanesi eksik."})
                fail_count += 1
                completed += 1
                await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
                continue
                
            # 2. Get file paths
            # GCB filename search
            beyan_path = None
            try:
                for f in os.listdir(beyan_dir):
                    if match_beyan_filename(gcb, f):
                        beyan_path = os.path.join(beyan_dir, f)
                        break
            except Exception as le:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] HATA: Beyanname klasörü okunamadı: {str(le)}", "level": "error"})
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": "Dizin okuma hatası."})
                fail_count += 1
                completed += 1
                await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
                continue
                        
            if not beyan_path:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] HATA: Beyanname PDF dosyası bulunamadı.", "level": "error"})
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": "Beyanname PDF eksik."})
                fail_count += 1
                completed += 1
                await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
                continue
                
            # Invoice files matching
            fatura_paths = []
            missing_invoices = []
            try:
                for fat in faturalar:
                    found_path = None
                    for f in os.listdir(fatura_dir):
                        if match_fatura_filename(fat, f):
                            found_path = os.path.join(fatura_dir, f)
                            break
                    if found_path:
                        fatura_paths.append(found_path)
                    else:
                        missing_invoices.append(fat)
            except Exception as le:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] HATA: Fatura klasörü okunamadı: {str(le)}", "level": "error"})
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": "Dizin okuma hatası."})
                fail_count += 1
                completed += 1
                await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
                continue
                    
            if missing_invoices:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] UYARI: Eksik fatura PDF'leri var: {missing_invoices}", "level": "warning"})
                
            if not fatura_paths:
                await websocket.send_json({"type": "merge_log", "message": f"[{gcb}] HATA: Eşleşen hiçbir fatura PDF'i bulunamadı. Atlattırılıyor.", "level": "error"})
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": "Faturalar eksik."})
                fail_count += 1
                completed += 1
                await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
                continue
                
            # 3. Perform PDF Merge
            output_path = os.path.join(output_dir, target_filename)
            
            def do_merge(b_path, f_paths, out_path):
                merger = PdfMerger()
                merger.append(b_path)
                for fp in f_paths:
                    merger.append(fp)
                merger.write(out_path)
                merger.close()
                
            try:
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, do_merge, beyan_path, fatura_paths, output_path)
                
                await websocket.send_json({
                    "type": "merge_log", 
                    "message": f"[{gcb}] BAŞARILI: {target_filename} dosyası oluşturuldu.", 
                    "level": "success"
                })
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "success", "output_name": target_filename})
                created_files.append((output_path, item.get("date", "")))
                success_count += 1
            except Exception as e:
                await websocket.send_json({
                    "type": "merge_log", 
                    "message": f"[{gcb}] HATA: PDF birleştirme başarısız: {str(e)}", 
                    "level": "error"
                })
                await websocket.send_json({"type": "merge_item_complete", "gcb": gcb, "status": "fail", "message": str(e)})
                fail_count += 1
                
            completed += 1
            await websocket.send_json({"type": "merge_progress", "completed": completed, "total": total})
            await asyncio.sleep(0.05)
            
        # Zip successfully merged documents
        zip_url = None
        if success_count > 0:
            await websocket.send_json({"type": "merge_log", "message": "[SİSTEM] Başarıyla birleştirilen belgeler ZIP arşivine sıkıştırılıyor...", "level": "info"})
            try:
                import zipfile
                def create_zip(z_path, files_info):
                    with zipfile.ZipFile(z_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for fp, date_str in files_info:
                            if os.path.exists(fp):
                                folder_name = ""
                                year, month = parse_date_to_year_month(date_str)
                                if year and month:
                                    months_tr = {
                                        "01": "Ocak", "02": "Subat", "03": "Mart", "04": "Nisan",
                                        "05": "Mayis", "06": "Haziran", "07": "Temmuz", "08": "Agustos",
                                        "09": "Eylul", "10": "Ekim", "11": "Kasim", "12": "Aralik"
                                    }
                                    month_name = months_tr.get(month, month)
                                    folder_name = f"{year}_{month}_{month_name}"
                                        
                                if not folder_name:
                                    folder_name = "Diger_Tarihsiz"
                                    
                                zip_path_in_archive = os.path.join(folder_name, os.path.basename(fp))
                                zipf.write(fp, zip_path_in_archive)
                
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, create_zip, old_zip, created_files)
                zip_url = f"/api/merge/download?session_id={session_id}"
                await websocket.send_json({"type": "merge_log", "message": "[SİSTEM] ZIP arşivi hazırlandı.", "level": "success"})
            except Exception as ze:
                await websocket.send_json({"type": "merge_log", "message": f"[UYARI] ZIP oluşturulamadı: {str(ze)}", "level": "warning"})
                
        await websocket.send_json({
            "type": "merge_finished",
            "success_count": success_count,
            "fail_count": fail_count,
            "zip_url": zip_url
        })
    except Exception as ge:
        try:
            await websocket.send_json({"type": "merge_log", "message": f"[HATA] Beklenmeyen sistem hatası: {str(ge)}", "level": "error"})
            await websocket.send_json({"type": "merge_finished", "success_count": success_count, "fail_count": fail_count + (total - completed), "zip_url": None})
        except Exception:
            pass
    finally:
        session.is_merge_running = False
        session.merge_task = None


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, session_id: str = None):
    session = get_session(session_id)
    await manager.connect(session_id, websocket)
    
    # Send initial state to the newly connected client
    res = read_excel_data(session.active_excel_path)
    session.gcb_col_idx = res["gcb_col_idx"]
    session.date_col_idx = res["date_col_idx"]
    session.fatura_col_idx = res["fatura_col_idx"]
    session.firma_col_idx = res["firma_col_idx"]
    
    await websocket.send_json({
        "type": "init_state",
        "is_running": session.is_running,
        "completed": session.completed_count,
        "total": session.total_count,
        "active_file": get_display_filename(session),
        "log_history": session.log_history,
        "data": res["rows"],
        "headers": res["headers"],
        "gcb_col_idx": session.gcb_col_idx,
        "date_col_idx": session.date_col_idx,
        "fatura_col_idx": session.fatura_col_idx,
        "firma_col_idx": session.firma_col_idx,
    })
    try:
        while True:
            data = await websocket.receive_text()
            payload = json.loads(data)
            action = payload.get("action")
            
            if action == "ping":
                await websocket.send_json({"type": "pong"})
                continue
                
            if action == "start_all":
                if session.is_running:
                    await websocket.send_json({"type": "log", "message": "Sorgulama zaten çalışıyor."})
                    continue
                
                session.bypass = payload.get("bypass", False)
                res = read_excel_data(session.active_excel_path)
                excel_rows = res["rows"]
                # Query EVERY row that has a GCB number (do not skip rows with existing intaç dates)
                pending = [r for r in excel_rows if r.get("gcb")]
                
                if not pending:
                    await websocket.send_json({"type": "log", "message": "Sorgulanacak beyanname numarası bulunamadı."})
                    await websocket.send_json({"type": "finished"})
                    continue
                
                await websocket.send_json({"type": "log", "message": f"Sorgulanacak {len(pending)} beyanname bulundu. İşlem başlatılıyor..."})
                await websocket.send_json({"type": "log", "message": f"[SİSTEM] 📌 Bulunan intaç tarihleri Excel dosyasının EN SAĞINDAKİ {session.date_col_idx}. Sütuna ('Sistem Gümrük İntaç Tarihi') yazılmaktadır."})
                session.task = asyncio.create_task(run_scraper_task(session_id, websocket, pending))
                
            elif action == "start_custom_list":
                if session.is_running:
                    await websocket.send_json({"type": "log", "message": "Sorgulama zaten çalışıyor."})
                    continue
                
                session.bypass = payload.get("bypass", False)
                raw_text = payload.get("raw_text", "").strip()
                if not raw_text:
                    await websocket.send_json({"type": "log", "message": "HATA: Gönderilen liste boş."})
                    continue
                
                lines = [line.strip() for line in raw_text.split("\n") if line.strip()]
                parsed_items = []
                
                for idx, line in enumerate(lines):
                    gcb, fatura, firma = parse_custom_line(line)
                    if gcb:
                        parsed_items.append({
                            "fatura": fatura or "",
                            "firma": firma or "",
                            "gcb": gcb
                        })
                    else:
                        await websocket.send_json({"type": "log", "message": f"[UYARI] Satır ayrıştırılamadı (Geçerli Beyanname No bulunamadı): '{line}'"})
                
                if not parsed_items:
                    await websocket.send_json({"type": "log", "message": "HATA: Geçerli hiçbir beyanname numarası ayrıştırılamadı."})
                    continue
                
                await websocket.send_json({"type": "log", "message": f"Ayrıştırma başarılı: {len(parsed_items)} adet beyanname bulundu."})
                
                # Generate new custom Excel with unique session id
                session_custom_path = os.path.join(BASE_DIR, f"EXPORT_CUSTOM_{session_id}.xlsx")
                generate_custom_excel(parsed_items, session_custom_path)
                session.active_excel_path = session_custom_path
                
                # Fetch fresh rows of the newly created custom Excel
                res = read_excel_data(session.active_excel_path)
                session.gcb_col_idx = res["gcb_col_idx"]
                session.date_col_idx = res["date_col_idx"]
                session.fatura_col_idx = res["fatura_col_idx"]
                session.firma_col_idx = res["firma_col_idx"]
                
                # Send rows back to update client UI
                await websocket.send_json({
                    "type": "custom_list_loaded",
                    "data": res["rows"],
                    "headers": res["headers"],
                    "gcb_col_idx": session.gcb_col_idx,
                    "date_col_idx": session.date_col_idx,
                    "fatura_col_idx": session.fatura_col_idx,
                    "firma_col_idx": session.firma_col_idx,
                    "active_file": get_display_filename(session)
                })
                
                await websocket.send_json({"type": "log", "message": "Yeni sorgu tablosu oluşturuldu. Headless sorgular başlatılıyor..."})
                session.task = asyncio.create_task(run_scraper_task(session_id, websocket, res["rows"]))
                
            elif action == "query_single":
                row_idx = payload.get("row")
                gcb = payload.get("gcb")
                if not row_idx or not gcb:
                    continue
                    
                if session.is_running:
                    await websocket.send_json({"type": "log", "message": "Arka planda çalışan bir sorgulama var, tekil sorgu yapılamaz."})
                    continue
                
                session.bypass = payload.get("bypass", False)
                await websocket.send_json({"type": "log", "message": f"Satır {row_idx} ({gcb}) için tekil sorgulama başlatılıyor..."})
                session.task = asyncio.create_task(run_scraper_task(session_id, websocket, [{"row": row_idx, "gcb": gcb}]))
                
            elif action == "reset_excel":
                if session.is_running:
                    await websocket.send_json({"type": "log", "message": "Sorgulama devam ederken tablo sıfırlanamaz."})
                    continue
                session_custom_path = os.path.join(BASE_DIR, f"EXPORT_CUSTOM_{session_id}.xlsx")
                session.active_excel_path = None
                if os.path.exists(session_custom_path):
                    try:
                        os.remove(session_custom_path)
                    except Exception:
                        pass
                await websocket.send_json({
                    "type": "custom_list_loaded",
                    "data": [],
                    "headers": [],
                    "gcb_col_idx": 9,
                    "date_col_idx": 12,
                    "fatura_col_idx": 1,
                    "firma_col_idx": 3,
                    "active_file": None
                })
                await websocket.send_json({"type": "log", "message": "[SİSTEM] Tablo sıfırlandı. Orijinal Excel bağlantısı kesildi. Yeni görev bekleniyor..."})
                
            elif action == "run_merge":
                if session.is_merge_running:
                    await websocket.send_json({"type": "merge_log", "message": "Evrak birleştirme işlemi zaten çalışıyor.", "level": "warning"})
                    continue
                    
                beyan_dir = payload.get("beyan_dir")
                fatura_dir = payload.get("fatura_dir")
                output_dir = payload.get("output_dir")
                items_to_merge = payload.get("items", [])
                
                if not items_to_merge:
                    await websocket.send_json({"type": "merge_log", "message": "HATA: Birleştirilecek geçerli bir veri listesi bulunamadı.", "level": "error"})
                    continue
                    
                session.merge_task = asyncio.create_task(run_pdf_merge_task(
                    session_id, websocket, beyan_dir, fatura_dir, output_dir, items_to_merge
                ))

            elif action == "stop":
                any_stopped = False
                if session.is_running:
                    session.is_running = False
                    session.cancel_event.set()  # Instant cancel signal
                    await websocket.send_json({"type": "log", "message": "Durdurma sinyali gönderildi — tüm işçiler durduruluyor..."})
                    any_stopped = True
                if session.is_merge_running:
                    session.is_merge_running = False
                    session.merge_cancel_event.set()
                    await websocket.send_json({"type": "merge_log", "message": "Durdurma sinyali gönderildi — evrak birleştirme sonlandırılıyor...", "level": "warning"})
                    any_stopped = True
                    
                if not any_stopped:
                    await websocket.send_json({"type": "log", "message": "Çalışan aktif bir işlem yok."})
                    
    except WebSocketDisconnect:
        manager.disconnect(session_id, websocket)
    except Exception as e:
        print("WebSocket Error:", e)
        manager.disconnect(session_id, websocket)

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    is_local = not os.environ.get("RENDER")
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=is_local)
